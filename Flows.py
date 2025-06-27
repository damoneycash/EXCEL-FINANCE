import pandas as pd
import matplotlib.pyplot as plt
import networkx as nx
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import numpy as np

# Liste des feuilles sur lesquelles ajouter l'image
sheets_to_update = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY","WEEKLY", "YESTERDAY"]

# Lecture du classeur Excel
file_path = "Weekly.xlsx"
wb = load_workbook(file_path)

# Parcourir chaque feuille et ajouter l'image
for sheet_name in sheets_to_update:
    # Supprimer l'image existante de la feuille
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for img in ws._images:
            ws._images.remove(img)

    # Lecture des données de la feuille spécifique
    df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=[1, 2, 3])

    # Supprimer les lignes avec des valeurs manquantes
    df = df.dropna()

    # Extraire les paires uniques sans doublons avec leurs performances
    pairs = {}
    for _, row in df.iterrows():
        pair = tuple([str(row[0]), str(row[1])])
        pairs[pair] = row[2]

    # Création du graphe orienté
    G = nx.DiGraph()

    # Ajout de chaque monnaie comme un nœud dans le graphe
    for pair in pairs.keys():
        for currency in pair:
            G.add_node(currency)

    # Ajout des arêtes entre chaque paire avec les performances comme poids
    for pair, performance in pairs.items():
        if abs(performance) >= 0.2:
            if performance >= 0.0:
                G.add_edge(pair[1], pair[0], weight=performance)
            else :
                G.add_edge(pair[0], pair[1], weight=performance)

    # Calcul de la largeur des liaisons en fonction de la performance en valeur absolue
    edge_width = [abs(performance) * 7 for performance in nx.get_edge_attributes(G, 'weight').values()]

    # Préparation des labels des liaisons en filtrant ceux dont la valeur absolue de la performance est inférieure ou égale à 0.3%
    edge_labels = nx.get_edge_attributes(G, 'weight')
    rounded_edge_labels = {(u, v): f"{abs(performance):.1f}%" for (u, v), performance in edge_labels.items() if abs(performance) > 4}

    # Définition de l'opacité des liaisons en fonction de la performance en valeur absolue
    edge_opacity = {edge: 1 - abs(performance) for edge, performance in nx.get_edge_attributes(G, 'weight').items()}

    # Affichage du graphe sous forme de cercle avec des liaisons de largeur variable et les labels filtrés
    plt.figure(figsize=(8, 8))
    pos = nx.circular_layout(G)

    node_colors = []
    for node in G.nodes():
        if node in ['CHF', 'JPY']:
            node_colors.append('green')
        elif node in ['AUD', 'NZD']:
            node_colors.append('red')
        elif node in ['GBP', 'EUR']:
            node_colors.append('orange')
        elif node in ['USD', 'CAD']:
            node_colors.append('brown')
        else:
            node_colors.append('skyblue')  # Default color


        # Définir les positions des nœuds pour que les couleurs similaires soient à côté
    color_order = ['green', 'red', 'orange', 'brown', 'skyblue']
    grouped_nodes = {color: [] for color in color_order}
    for node, color in zip(G.nodes(), node_colors):
        grouped_nodes[color].append(node)

    # Flatten the grouped nodes list while maintaining the order of colors
    ordered_nodes = [node for color in color_order for node in grouped_nodes[color]]

    # Define the circular layout with ordered nodes
    pos = nx.circular_layout(G)
    angle = 2 * np.pi / 8
    for i, node in enumerate(ordered_nodes):
        pos[node] = (np.cos(i * angle), np.sin(i * angle))

    # Dessin des liaisons avec des couleurs spécifiques en fonction de la destination et de l'opacité en fonction de la performance
    edge_colors = []
    for edge in G.edges():
        source, target = edge
        if (target in ['JPY', 'CHF'] and source in ['JPY', 'CHF']):
            edge_colors.append('green')
        elif source in ['JPY', 'CHF']:
            edge_colors.append('red')
        elif target in ['JPY', 'CHF']:
            edge_colors.append('green')
        elif source in ['NZD', 'AUD']:
            edge_colors.append('green')
        elif target in ['NZD', 'AUD']:
            edge_colors.append('red')
        elif source in ['EUR', 'GBP'] and target in ['USD', 'EUR', 'CAD', 'GBP']:
            edge_colors.append('blue')
        elif source in ['USD', 'CAD'] and target in ['USD', 'EUR', 'CAD', 'GBP']:
            edge_colors.append('black')
        else:
            edge_colors.append('black')

    nx.draw(G, pos, with_labels=True, node_size=2000, node_color=node_colors, font_size=10, font_weight="bold", width=edge_width, edge_color=edge_colors, arrows=True)
    nx.draw_networkx_edge_labels(G, pos, edge_labels=rounded_edge_labels, font_size=8, alpha=0.8)
    plt.figtext(0.5, 0.01, "© Damoneycash", ha="center", fontsize=10)
    # Enregistrement de l'image dans un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)

    # Ajout de l'image sur la feuille spécifiée
    ws = wb[sheet_name]
    img = Image(buffer)
    ws.add_image(img, 'F1')  # Position de l'image sur chaque feuille, cellule F1

    # Réinitialisation de la figure pour la prochaine feuille
    plt.close()

# Sauvegarde du classeur Excel avec l'image ajoutée sur chaque feuille
wb.save(file_path)


